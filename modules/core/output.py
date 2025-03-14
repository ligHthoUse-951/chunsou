# -*- coding: utf-8 -*-

'''
@File    ：output.py
@IDE     ：PyCharm
@Author  ：Funsiooo
@Github  ：https://github.com/Funsiooo
'''


from modules.core.time import print_start_time,current_time
from modules.core.color import Colors
from modules.core.args import argument
import csv
from openpyxl import Workbook
import os
from openpyxl.styles import Font, PatternFill, Border, Side
from modules.core.check import check_version


def script_start():
    # 脚本启动时输出
    # 检查程序版本
    check_version()
    print(f"{Colors.CYAN}{print_start_time()} {Colors.GREEN}[*]{Colors.RESET} {Colors.GREEN}[INFO"
          f"]{Colors.RESET} {Colors.GREEN}fingerprint recognition result order | Website fingerprint | Web title | "
          f"Web stack{Colors.RESET}")
    # print(f"{Colors.CYAN}{print_start_time()} {Colors.GREEN}[*]{Colors.RESET} {Colors.GREEN}[INFO"
    #       f"]{Colors.RESET} {Colors.GREEN}the program starts running, and the following are the fingerprint recognition results{Colors.RESET}")


def script_end():
    # 执行脚本结束提示语
    print(f"{Colors.CYAN}{print_start_time()} {Colors.GREEN}[*]{Colors.RESET} {Colors.GREEN}[INFO"
          f"]{Colors.RESET} {Colors.YELLOW}{Colors.GREEN}the script has finished "
          f"executing, and the scan results are saved in the{Colors.RESET} {Colors.BROWN}{output_dir()}{Colors.RESET}"
          f" , {Colors.GREEN}scan end @ {current_time()}{Colors.RESET}")


def tip_start():
    # 脚本启动时输出
    print(f"{Colors.CYAN}{print_start_time()} {Colors.GREEN}[*]{Colors.RESET} {Colors.GREEN}[INFO"
          f"]{Colors.RESET} {Colors.GREEN} The basic syntax for mapping in FOFA and Hunter spaces is as follows. Please be mindful of user permissions when using it {Colors.RESET}")
    # print(f"{Colors.CYAN}{print_start_time()} {Colors.GREEN}[*]{Colors.RESET} {Colors.GREEN}[INFO"
    #       f"]{Colors.RESET} {Colors.GREEN}the program starts running, and the following are the fingerprint recognition results{Colors.RESET}")


def output_dir():
    args = argument()
    if args.output:
        out_file = args.output
        output_dir = out_file
        return output_dir

    else:
        out_file = 'results/results.txt'
        output_dir = out_file
        return output_dir


def fofa_output_dir():
    args = argument()
    if args.output:
        out_file = args.output
        return out_file
    else:
        out_file = 'results/fofa_result.xlsx'
        return out_file


def hunter_output_dir():
    args = argument()
    if args.output:
        out_file = args.output
        return out_file
    else:
        out_file = 'results/hunter_result.xlsx'
        return out_file


def fofa_start():
    # 脚本启动时输出
    # 检查程序版本
    check_version()
    print(f"{Colors.CYAN}{print_start_time()} {Colors.GREEN}[*]{Colors.RESET} {Colors.GREEN}[INFO"
          f"]{Colors.RESET} {Colors.GREEN}to start calling the hunter api, you need to correctly configure the api in the modules/config/config.ini file{Colors.RESET}")


def fofa_end():
    # 执行脚本结束提示语
    print(f"{Colors.CYAN}{print_start_time()} {Colors.GREEN}[*]{Colors.RESET} {Colors.GREEN}[INFO"
          f"]{Colors.RESET} {Colors.YELLOW}{Colors.GREEN}the script has finished "
          f"executing, and the scan results are saved in the{Colors.RESET} {Colors.BROWN}{fofa_output_dir()}{Colors.RESET}"
          f" , {Colors.GREEN}scan end @ {current_time()}{Colors.RESET}")


def hunter_start():
    # 脚本启动时输出
    # 检查程序版本
    check_version()
    print(f"{Colors.CYAN}{print_start_time()} {Colors.GREEN}[*]{Colors.RESET} {Colors.GREEN}[INFO"
          f"]{Colors.RESET} {Colors.GREEN}to start calling the hunter api, you need to correctly configure the api in the modules/config/config.ini file{Colors.RESET}")


def hunter_end():
    # 执行脚本结束提示语
    print(f"{Colors.CYAN}{print_start_time()} {Colors.GREEN}[*]{Colors.RESET} {Colors.GREEN}[INFO"
          f"]{Colors.RESET} {Colors.YELLOW}{Colors.GREEN}the script has finished "
          f"executing, and the scan results are saved in the{Colors.RESET} {Colors.BROWN}{hunter_output_dir()}{Colors.RESET}"
          f" , {Colors.GREEN}scan end @ {current_time()}{Colors.RESET}")


# 指纹识别保存为xlsx
def scan_save_to_excle(results):
    out_file = output_dir()
    csv_headers = ['状态码', '网页URL', '网页标题', '网站技术栈', '指纹结果']
    filename = 'output.csv'
    with open(filename, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(csv_headers)

    wb = Workbook()
    ws = wb.active

    title_font = Font(bold=True, name='宋体', size=12)
    title_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")  # 设置背景颜色为黄色

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))
    for col_num, header in enumerate(csv_headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = title_font
        cell.fill = title_fill
        cell.border = thin_border

    content_font = Font(name='Calibri', size=11)  # 设置内容字体为Calibri，大小为11
    for row_num, result in enumerate(results, 2):  # 从第二行开始添加内容
        for col_num, value in enumerate(result, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.font = content_font

            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                 bottom=Side(style='thin'))
            cell.border = thin_border

    for column in ws.columns:
        max_length = 0
        for cell in column:
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        adjusted_width = (max_length + 1) * 1.2
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    adjusted_filename = out_file
    wb.save(adjusted_filename)

    os.remove(filename)

# fofa 输出结果
def fofa_save_to_excel(results):
    out_file = fofa_output_dir()

    csv_headers = ['Host', '网页标题', '域名', 'Link','IP', 'Port', 'Protocol', 'Server']
    filename = 'output.csv'
    with open(filename, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(csv_headers)

    wb = Workbook()
    ws = wb.active

    title_font = Font(bold=True, name='宋体', size=12)
    title_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")  # 设置背景颜色为黄色
    for col_num, header in enumerate(csv_headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = title_font
        cell.fill = title_fill

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))
        cell.border = thin_border

    content_font = Font(name='Calibri', size=11)  # 设置内容字体为Calibri，大小为11
    for row_num, result in enumerate(results, 2):  # 从第二行开始添加内容
        for col_num, value in enumerate(result, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.font = content_font

            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            cell.border = thin_border

    for column in ws.columns:
        max_length = 0
        for cell in column:
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        adjusted_width = (max_length + 1) * 1.2
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    adjusted_filename = out_file
    wb.save(adjusted_filename)

    os.remove(filename)


def hunter_save_to_excel(results):
    out_file = hunter_output_dir()
    csv_headers = ['网址', 'IP', '端口', '网站标题', '域名', '状态码', '协议名称', '基础协议', '系统名称', '公司名称', '备案号', '国家', '省份','应用名称','应用版本']
    filename = 'output.csv'
    with open(filename, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(csv_headers)

    wb = Workbook()
    ws = wb.active

    title_font = Font(bold=True, name='宋体', size=12)
    title_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")  # 设置背景颜色为黑灰色

    for col_num, header in enumerate(csv_headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = title_font
        cell.fill = title_fill

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))
        cell.border = thin_border


    content_font = Font(name='Calibri', size=11)  # 设置内容字体为Calibri，大小为11
    for result in results:
        row_values = list(result.values())
        ws.append(row_values)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = content_font
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                 bottom=Side(style='thin'))
            cell.border = thin_border

    for column in ws.columns:
        max_length = 0
        for cell in column:
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        adjusted_width = (max_length + 2) * 1.2  # 调整列宽的公式，可以根据需要进行调整
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    adjusted_filename = out_file
    wb.save(adjusted_filename)

    os.remove(filename)